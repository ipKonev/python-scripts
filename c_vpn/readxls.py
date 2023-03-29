from openpyxl import load_workbook
from pprint import pprint

def read_xlsx(name='files/res.xlsx',dns='files/dns_list.xlsx'):
    hosts=[]
    dns_xlsx=load_workbook(filename=dns)
    dns_list=dns_xlsx['addr_info']
    wb=load_workbook(filename=name)
    sr=wb['addr_info']
    vpn_name=sr['A1'].value
    j=2
    while dns_list[f'B{j}'].value != None:
        host=(dns_list[f'B{j}'].value,dns_list[f'C{j}'].value,dns_list[f'D{j}'].value,dns_list[f'E{j}'].value)
        hosts.append(host)
        j+=1

    i=2
    while sr[f'B{i}'].value != None:
        host=(sr[f'B{i}'].value,sr[f'C{i}'].value,sr[f'D{i}'].value,sr[f'E{i}'].value)
        hosts.append(host)
        i+=1
    return vpn_name,hosts
            
if __name__ == '__main__':
    a=read_xlsx('files/res.xlsx')
    pprint(create_ace(a))
